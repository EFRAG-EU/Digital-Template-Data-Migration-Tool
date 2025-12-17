from openpyxl import load_workbook
import pandas as pd
from tkinter import Tk, filedialog
import warnings
import time

from outils.apply_changes_NR import apply_changes_NR, change_wastes
from outils.clean_NR_with_no_data import clean_NR_with_no_data
from outils.paste_values import paste_values
from outils.access_NR_tables import access_NR_table, access_missingNR_table
from outils.copy_values import copy_values
from outils.classes import values
from pickles.missingNR_df import missingNR_df

mapping_wastes = pd.read_pickle("pickles/Mapping_wastes.pkl")

list_migrationissues = []


def flatten_sublists_lc(nested_list):
    return [
        item for sublist in nested_list for item in sublist
    ]  # to later flatten list of issues


def tool(
    file_path: str,
    template_path: str = "Template/VSME-Digital-Template-1.1.1.xlsx",
):
    """
    Run the migration tool on the given Excel file_path. If file_path is None,
    a file dialog will be shown to pick the file.

    Returns:
        saved_path (str): path to the saved migrated template.
        elapsed (float): execution time in seconds.

    Raises:
        RuntimeError: on user cancellation or invalid template.
    """
    start_time = time.time()

    # Ask for file if not provided
    if not file_path:
        root = Tk()
        root.withdraw()
        file_path = filedialog.askopenfilename(
            title="Select Your Old Template Version (with data, .xlsx format)",
            filetypes=[("Excel Files", "*.xlsx")],
        )
        root.destroy()

    if not file_path:
        raise RuntimeError("No file selected.")

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", category=UserWarning)
        old_wb = load_workbook(file_path)

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", category=UserWarning)
        new_wb_empty = load_workbook(template_path)

    version_cell = old_wb["Introduction"].cell(row=1, column=3).value
    version_cell_new = new_wb_empty["Introduction"].cell(row=1, column=3).value

    df_old = access_NR_table(old_wb.defined_names)
    df_new = access_NR_table(new_wb_empty.defined_names)

    missingNR_df_old = access_missingNR_table(missingNR_df, version_cell)
    missingNR_df_new = access_missingNR_table(missingNR_df, version_cell_new)

    df_old_wv = copy_values(old_wb, df_old, key="name_ranges")
    missingNR_df_old_values = copy_values(old_wb, missingNR_df_old, key=None)

    if version_cell == "1.0.0":
        for position in [0, 1, 6]:
            missingNR_df_old_values[position].convert_month_to_numbers()

    df_old_tomerge = apply_changes_NR(df_old_wv, version_cell, version_cell_new)
    df_old_tomerge = clean_NR_with_no_data(df_old_tomerge)
    if version_cell in ["1.0.0", "1.0.1", "1.1.0"]:
        list_migrationissues.append(change_wastes(df_old_tomerge, mapping_wastes))

    df_new_wv = df_new.merge(df_old_tomerge)
    missingNR_df_new_wv = pd.concat([missingNR_df_new, missingNR_df_old_values], axis=1)

    if version_cell in ["1.0.0", "1.0.1"]:
        length = (
            df_new_wv.loc[
                df_new_wv["name_ranges"] == "CountryOfEmploymentContractAxis",
                "cell_values",
            ]
            .values[0]
            .count_uniques()
        )
        if length > 2:  # ignore PyLance warning, no issue at runtime
            missingNR_df_new_wv.loc[
                (missingNR_df_new_wv["sheets"] == "Social Disclosures")
                & (missingNR_df_new_wv["cell_ranges"] == "$E$27"),
                "cell_values",
            ] = values([[True]])
        else:
            missingNR_df_new_wv.loc[
                (missingNR_df_new_wv["sheets"] == "Social Disclosures")
                & (missingNR_df_new_wv["cell_ranges"] == "$E$27"),
                "cell_values",
            ] = values([[False]])

    paste_values(new_wb_empty, missingNR_df_new_wv)
    paste_values(new_wb_empty, df_new_wv, NR=True)

    saved_path = f"Template/result_from_{version_cell}.xlsx"
    new_wb_empty.save(saved_path)

    elapsed = time.time() - start_time

    return new_wb_empty, elapsed, flatten_sublists_lc(list_migrationissues)


tool(
    "Template/VSME-Digital-Template-Sample-1.0.0.xlsx",
    "Template/VSME-Digital-Template-1.1.1.xlsx",
)
