"""Determinism tests: repeated migrations must produce identical results.

These tests run the migration multiple times per sample so are inherently
slow.  They are marked with ``@pytest.mark.slow`` and can be skipped in
fast CI runs with::

    pytest -m "not slow"
"""

from __future__ import annotations

from pathlib import Path
from typing import NamedTuple

import pytest
from conftest import SAMPLE_PARAMS
from openpyxl import Workbook
from openpyxl.worksheet.formula import ArrayFormula

from migration_tool.tool import migrate_workbook

pytestmark = pytest.mark.slow


class TwoRunResult(NamedTuple):
    """Outcome of running migration twice on the same sample."""

    sample_path: Path
    wb_a: Workbook
    wb_b: Workbook
    issues_a: list[str]
    issues_b: list[str]


class TestDeterminism:
    """Migration must be deterministic: same input → same output."""

    @pytest.fixture(params=SAMPLE_PARAMS, scope="module")
    def sample_path(self, request):
        return request.param

    @pytest.fixture(scope="module")
    def two_runs(self, sample_path) -> TwoRunResult:
        """Run migrate_workbook twice on the same input."""
        wb_a, _, issues_a = migrate_workbook(sample_path)
        wb_b, _, issues_b = migrate_workbook(sample_path)
        return TwoRunResult(
            sample_path=sample_path,
            wb_a=wb_a,
            wb_b=wb_b,
            issues_a=issues_a,
            issues_b=issues_b,
        )

    def test_issues_are_identical(self, two_runs: TwoRunResult):
        assert two_runs.issues_a == two_runs.issues_b, (
            f"Non-deterministic issues for {two_runs.sample_path.name}\n"
            f"  Run 1: {two_runs.issues_a}\n"
            f"  Run 2: {two_runs.issues_b}"
        )

    def test_cell_values_are_identical(self, two_runs: TwoRunResult):
        """Every cell in every sheet must have the same value across runs."""
        wb_a = two_runs.wb_a
        wb_b = two_runs.wb_b
        name = two_runs.sample_path.name

        assert wb_a.sheetnames == wb_b.sheetnames, (
            f"Sheet names differ for {name}: {wb_a.sheetnames} vs {wb_b.sheetnames}"
        )

        for sheet_name in wb_a.sheetnames:
            ws_a = wb_a[sheet_name]
            ws_b = wb_b[sheet_name]

            assert ws_a.max_row == ws_b.max_row, (
                f"Row count differs in sheet '{sheet_name}' for {name}"
            )
            assert ws_a.max_column == ws_b.max_column, (
                f"Column count differs in sheet '{sheet_name}' for {name}"
            )

            for row_idx in range(1, (ws_a.max_row or 0) + 1):
                for col_idx in range(1, (ws_a.max_column or 0) + 1):
                    val_a = _comparable(ws_a.cell(row=row_idx, column=col_idx).value)
                    val_b = _comparable(ws_b.cell(row=row_idx, column=col_idx).value)
                    assert val_a == val_b, (
                        f"Cell ({row_idx}, {col_idx}) differs in sheet "
                        f"'{sheet_name}' for {name}: {val_a!r} vs {val_b!r}"
                    )


def _comparable(value: object) -> object:
    """Normalise openpyxl rich objects to plain comparable representations.

    Some cell values (e.g. ``ArrayFormula``) are rich objects that compare
    by identity.  Convert them to a tuple of their meaningful attributes so
    equality checks work on *content* rather than memory address.
    """
    if isinstance(value, ArrayFormula):
        return ("ArrayFormula", value.text, value.ref)
    # Add further cases here as they surface, e.g.:
    #   from openpyxl.worksheet.merge import MergedCell
    #   if isinstance(value, SomeOtherType): ...
    return value
