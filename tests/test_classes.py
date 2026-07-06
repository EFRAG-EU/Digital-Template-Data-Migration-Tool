"""Unit tests for the Shape/Value helper classes in migration_tool.classes.

These pin down the behaviour-preserving refactor that introduces NullShape /
NullValue subclasses in place of scattered ``is not None`` guards.
"""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from migration_tool.classes import (
    NullShape,
    NullValue,
    Shape,
    Value,
    check_formula,
    make_shape,
    make_value,
)


@pytest.fixture
def sheet():
    """A worksheet with a 3x3 block of values and one formula cell."""
    wb = Workbook()
    ws = wb.active
    for row in range(1, 4):
        for col in range(1, 4):
            ws.cell(row=row, column=col).value = row * 10 + col
    return ws


# --- check_formula ---------------------------------------------------------


def test_check_formula_detects_formula():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "=1+1"
    ws["A2"] = 42
    assert check_formula(ws["A1"]) is True
    assert check_formula(ws["A2"]) is False


# --- Shape -----------------------------------------------------------------


class TestShape:
    def test_geometry_getters(self):
        shape = Shape((1, 2, 4, 6))  # left, top, right, bottom
        assert shape.left() == 1
        assert shape.top() == 2
        assert shape.right() == 4
        assert shape.bottom() == 6

    def test_rows_and_cols(self):
        shape = Shape((1, 1, 3, 4))
        assert shape.cols() == 3  # 1..3
        assert shape.rows() == 4  # 1..4

    def test_isonecell_true(self):
        assert Shape((2, 2, 2, 2)).isonecell() is True

    def test_isonecell_false(self):
        assert Shape((1, 1, 2, 2)).isonecell() is False

    def test_truthiness(self):
        assert bool(Shape((1, 1, 1, 1))) is True

    def test_build_values_single_cell(self, sheet):
        shape = Shape((1, 1, 1, 1))
        assert shape.build_values(sheet) == [[11]]

    def test_build_values_block(self, sheet):
        shape = Shape((1, 1, 3, 3))
        assert shape.build_values(sheet) == [
            [11, 12, 13],
            [21, 22, 23],
            [31, 32, 33],
        ]

    def test_build_values_single_formula_cell_returns_none(self, sheet):
        sheet.cell(row=1, column=1).value = "=2+2"
        assert Shape((1, 1, 1, 1)).build_values(sheet) is None

    def test_build_values_block_with_formula_returns_none(self, sheet):
        sheet.cell(row=2, column=2).value = "=2+2"
        assert Shape((1, 1, 3, 3)).build_values(sheet) is None


# --- NullShape -------------------------------------------------------------


class TestNullShape:
    def test_is_a_shape(self):
        assert isinstance(NullShape(), Shape)

    def test_falsy(self):
        assert bool(NullShape()) is False

    def test_isonecell_false(self):
        assert NullShape().isonecell() is False

    def test_build_values_none(self, sheet):
        assert NullShape().build_values(sheet) is None

    @pytest.mark.parametrize(
        "method", ["left", "top", "right", "bottom", "rows", "cols"]
    )
    def test_geometry_raises(self, method):
        with pytest.raises(NotImplementedError):
            getattr(NullShape(), method)()


# --- make_shape ------------------------------------------------------------


class TestMakeShape:
    def test_tuple_makes_shape(self):
        shape = make_shape((1, 1, 2, 2))
        assert type(shape) is Shape
        assert shape.left() == 1

    def test_range_string_makes_shape(self):
        # "B2:D4" -> left=2, top=2, right=4, bottom=4
        shape = make_shape("B2:D4")
        assert type(shape) is Shape
        assert (shape.left(), shape.top(), shape.right(), shape.bottom()) == (
            2,
            2,
            4,
            4,
        )

    def test_single_cell_range_string(self):
        assert make_shape("A1").isonecell() is True

    def test_none_makes_null_shape(self):
        assert type(make_shape(None)) is NullShape

    def test_empty_string_makes_null_shape(self):
        assert type(make_shape("")) is NullShape

    def test_open_range_raises_value_error(self):
        # A whole-column range has no fixed top/bottom -> range_boundaries
        # returns None for those dimensions.
        with pytest.raises(ValueError):
            make_shape("A:A")


# --- Value -----------------------------------------------------------------


class TestValue:
    def test_values(self):
        block = [[1, 2], [3, 4]]
        assert Value(block).values() == block

    def test_topleft(self):
        assert Value([[7, 8], [9, 10]]).topleft() == 7

    def test_height_and_width(self):
        value = Value([[1, 2, 3], [4, 5, 6]])
        assert value.height() == 2
        assert value.width() == 3

    def test_truthiness(self):
        assert bool(Value([[1]])) is True

    def test_count_uniques(self):
        assert Value([[1, 2], [2, 3]]).count_uniques() == 3

    def test_add_checkboxes_turns_leading_none_into_false(self):
        value = Value([[None], [True], [None]])
        value.add_checkboxes()
        assert value.values() == [[False], [True], [False]]

    def test_convert_month_to_numbers(self):
        # second row holds the month (see template)
        value = Value([["header"], ["March"]])
        value.convert_month_to_numbers()
        assert value.values()[1][0] == 3

    def test_first_element_row_flat(self):
        value = Value([[1, 2, 3], [4, 5, 6]])
        assert value.first_element_row(double_list=False) == [1, 4]

    def test_first_element_row_mutates_in_place(self):
        value = Value([[1, 2, 3], [4, 5, 6]])
        value.first_element_row()
        assert value.values() == [[1], [4]]

    def test_paste_single_column(self):
        wb = Workbook()
        ws = wb.active
        value = Value([[1], [2], [3]])
        value.paste(ws, Shape((1, 1, 1, 3)))
        assert [ws.cell(row=r, column=1).value for r in range(1, 4)] == [1, 2, 3]

    def test_paste_block(self):
        wb = Workbook()
        ws = wb.active
        value = Value([[1, 2], [3, 4]])
        value.paste(ws, Shape((1, 1, 2, 2)))
        assert ws.cell(row=1, column=1).value == 1
        assert ws.cell(row=2, column=2).value == 4

    def test_enlarged_range_correction_pads_rows(self):
        value = Value([[1], [2]])
        value.enlarged_range_correction(Shape((1, 1, 1, 4)))  # 4 rows expected
        assert value.values() == [[1], [2], [None], [None]]


# --- NullValue -------------------------------------------------------------


class TestNullValue:
    def test_is_a_value(self):
        assert isinstance(NullValue(), Value)

    def test_values_none(self):
        assert NullValue().values() is None

    def test_topleft_none(self):
        assert NullValue().topleft() is None

    def test_count_uniques_zero(self):
        assert NullValue().count_uniques() == 0

    def test_first_element_row_empty(self):
        assert NullValue().first_element_row(double_list=False) == []

    def test_mutators_are_noops(self):
        value = NullValue()
        # none of these should raise
        value.add_checkboxes()
        value.convert_month_to_numbers()
        value.first_element_row()
        value.enlarged_range_correction(NullShape())
        value.paste(Workbook().active, NullShape())
        assert value.values() is None


# --- make_value ------------------------------------------------------------


class TestMakeValue:
    def test_block_makes_value(self):
        value = make_value([[1, 2]])
        assert type(value) is Value
        assert value.values() == [[1, 2]]

    def test_none_makes_null_value(self):
        assert type(make_value(None)) is NullValue
