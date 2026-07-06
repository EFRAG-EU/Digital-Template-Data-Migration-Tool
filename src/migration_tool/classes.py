from __future__ import annotations

from datetime import datetime
from typing import TypeAlias

from openpyxl.cell import Cell, MergedCell, ReadOnlyCell
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

CellType: TypeAlias = Cell | MergedCell | ReadOnlyCell
CellValue: TypeAlias = str | int | float | bool | datetime | None
# A single cell value as openpyxl hands them back, and a rectangular block of them.
Block: TypeAlias = list[list[CellValue]]


def check_formula(cell: CellType) -> bool:
    """Check whether a cell contains a formula."""
    return cell.data_type == "f"


class Shape:
    """A rectangular cell range (left, top, right, bottom).

    Backed by a real 4-tuple of 1-based coordinates. The "no shape" case is
    handled by :class:`NullShape` rather than by ``None`` checks in every method;
    use :func:`make_shape` to build the right one from an optional tuple.
    """

    def __init__(self, boundaries: tuple[int, int, int, int]):
        self._left, self._top, self._right, self._bottom = boundaries

    # Getters
    def left(self) -> int:
        return self._left

    def top(self) -> int:
        return self._top

    def right(self) -> int:
        return self._right

    def bottom(self) -> int:
        return self._bottom

    def rows(self) -> int:
        return self._bottom - self._top + 1

    def cols(self) -> int:
        return self._right - self._left + 1

    def isonecell(self) -> bool:
        return self.rows() == 1 and self.cols() == 1

    def build_values(self, sheet: Worksheet) -> Block | None:
        """Read the cell values inside this shape as a list of rows.

        Returns ``None`` if any cell in the range holds a formula (formula cells
        are recomputed when the new workbook is opened, so we must not copy them).
        """
        if self.isonecell():
            topleft = sheet.cell(row=self._top, column=self._left)
            if check_formula(topleft):
                return None
            return [[topleft.value]]

        list_values: Block = []
        for row in range(self._top, self._bottom + 1):
            row_values: list[CellValue] = []
            for col in range(self._left, self._right + 1):
                cell = sheet.cell(row=row, column=col)
                if check_formula(cell):
                    return None
                row_values.append(cell.value)
            list_values.append(row_values)
        return list_values

    def __bool__(self) -> bool:
        return True


class NullShape(Shape):
    """Sentinel shape for a name range with no readable destination.

    Falsy, never "one cell", and yields no values. Its geometry is undefined, so
    asking for it raises rather than silently returning ``None``.
    """

    def __init__(self):
        pass

    def __bool__(self) -> bool:
        return False

    def isonecell(self) -> bool:
        return False

    def build_values(self, sheet: Worksheet) -> None:
        return None

    def left(self) -> int:
        raise NotImplementedError("NullShape has no geometry")

    def top(self) -> int:
        raise NotImplementedError("NullShape has no geometry")

    def right(self) -> int:
        raise NotImplementedError("NullShape has no geometry")

    def bottom(self) -> int:
        raise NotImplementedError("NullShape has no geometry")

    def rows(self) -> int:
        raise NotImplementedError("NullShape has no geometry")

    def cols(self) -> int:
        raise NotImplementedError("NullShape has no geometry")


def make_shape(value: str | tuple[int, int, int, int] | None) -> Shape:
    """Build a :class:`Shape` from an Excel range string (e.g. ``"B2:D4"``) or a
    ready-made coordinate tuple; ``None`` or an empty range yields a
    :class:`NullShape`.

    Raises :class:`ValueError` if a range string has no fixed bounds (e.g. a
    whole-column ``"A:A"`` reference, where ``range_boundaries`` returns ``None``
    for the open dimensions)."""
    match value:
        case None | "" | ():
            return NullShape()
        case str():
            left, top, right, bottom = range_boundaries(value)
            if left is None or top is None or right is None or bottom is None:
                raise ValueError(f"Range {value!r} does not have fixed boundaries")
            return Shape((left, top, right, bottom))
        case _:
            return Shape(value)


class Value:
    """A rectangular block of cell values (e.g. ``[[1, 2], [3, 4]]``).

    The "no values" case (e.g. the source held a formula) is handled by
    :class:`NullValue` rather than by ``None`` checks in every method; use
    :func:`make_value` to build the right one from an optional block.
    """

    def __init__(self, block: Block):
        self._block = block

    def values(self) -> Block | None:
        return self._block

    def height(self) -> int:
        return len(self._block)

    def width(self) -> int:
        return len(self._block[0])

    def topleft(self) -> CellValue:
        return self._block[0][0]

    def first_element_row(self, double_list: bool = True) -> Block | list[CellValue]:
        """Keep only the first element of each row.

        With ``double_list`` (the default) the block is reduced in place to a
        single column; otherwise the first column is returned as a flat list.
        """
        if double_list:
            for row in range(self.height()):
                self._block[row] = [self._block[row][0]]
            return self._block
        return [self._block[row][0] for row in range(self.height())]

    def enlarged_range_correction(self, shape: Shape) -> None:
        """Pad the block with empty rows so it fills an enlarged shape."""
        if shape and self.height() != shape.rows():
            for _ in range(shape.rows() - self.height()):
                self._block.append([None])

    def add_checkboxes(self) -> None:
        """Treat a missing first-column value as an unticked checkbox (``False``)."""
        for row in range(self.height()):
            if self._block[row][0] is None:
                self._block[row][0] = False

    def convert_month_to_numbers(self) -> None:
        month_dict = {
            "January": 1,
            "February": 2,
            "March": 3,
            "April": 4,
            "May": 5,
            "June": 6,
            "July": 7,
            "August": 8,
            "September": 9,
            "October": 10,
            "November": 11,
            "December": 12,
        }
        month = self._block[1][0]  # second row of each element (see the Template)
        if month:
            self._block[1][0] = month_dict[month]

    def count_uniques(self) -> int:
        return len({item for sublist in self._block for item in sublist})

    def paste(self, sheet: Worksheet, shape: Shape) -> None:
        if not shape:
            return
        if self.width() == 1:
            for row in range(shape.top(), shape.bottom() + 1):
                sheet.cell(row=row, column=shape.left()).value = self._block[
                    row - shape.top()
                ][0]
        else:
            for row in range(shape.top(), shape.bottom() + 1):
                for col in range(shape.left(), shape.right() + 1):
                    sheet.cell(row=row, column=col).value = self._block[
                        row - shape.top()
                    ][col - shape.left()]

    def __bool__(self) -> bool:
        return True


class NullValue(Value):
    """Sentinel value for a name range with nothing to copy (e.g. a formula).

    Falsy, reports no values, and all mutators / paste are safe no-ops.
    """

    def __init__(self):
        pass

    def __bool__(self) -> bool:
        return False

    def values(self) -> None:
        return None

    def topleft(self) -> None:
        return None

    def first_element_row(self, double_list: bool = True) -> list[CellValue]:
        return []

    def enlarged_range_correction(self, shape: Shape) -> None:
        return None

    def add_checkboxes(self) -> None:
        return None

    def convert_month_to_numbers(self) -> None:
        return None

    def count_uniques(self) -> int:
        return 0

    def paste(self, sheet: Worksheet, shape: Shape) -> None:
        return None


def make_value(block: Block | None) -> Value:
    """Build a :class:`Value` from a block, or a :class:`NullValue`."""
    return Value(block) if block is not None else NullValue()
