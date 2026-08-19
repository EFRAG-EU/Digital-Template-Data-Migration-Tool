from typing import Dict, overload, Literal

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import absolute_coordinate, quote_sheetname, range_boundaries
from openpyxl.workbook.defined_name import DefinedName

from .classes import Shape, check_formula, make_shape, make_value
from .data.validations import VersionCollection


def get_version(wb: Workbook) -> str:

    versionNR = "template_reporting_template_version"

    try:
        sheet, rng = list(wb.defined_names[versionNR].destinations)[0]
    except KeyError:
        print("Name range for template version has changed.")

    shape = Shape(rng.range_boundaries())
    if not isinstance(val := wb[sheet].cell(shape._left, shape._top).value, str):
        raise ValueError(
            f"Expected str in version cell, got {type(val).__name__}: {val!r}"
        )

    return val


def check_status_incomplete(openpyxl_obj) -> bool:
    """Check if the workbook is filled out or not based on the value of the 'Status' cell in the 'Table of Contents & Validation' sheet"""

    status_cell = openpyxl_obj["Table of Contents & Validation"]["C3"].value
    "INCOMPLETE, INCOMPLETE, UFÆRDIG , ONVOLLEDIG, INCOMPLET, UNVOLLSTÄNDIG, NEAMHIOMLÁN, INCOMPLETO, NEBAIGTA, NIEKOMPLETNY, INCOMPLETO, NEPOPOLNO, INCOMPLETO"
    if status_cell in [
        "INCOMPLETE",
        "INCOMPLETE",
        "UFÆRDIG",
        "ONVOLLEDIG",
        "INCOMPLET",
        "UNVOLLSTÄNDIG",
        "NEAMHIOMLÁN",
        "INCOMPLETO",
        "NEBAIGTA",
        "NIEKOMPLETNY",
        "INCOMPLETO",
        "NEPOPOLNO",
        "INCOMPLETO",
    ]:
        return True
    else:
        return False


def create_table_of_contents(wb_values) -> Dict[str, int]:
    """Create a dict with keys as names of ToC sections and values as the corresponding row numbers in 'Table of Contents & Validation' sheet.
    If there is any change in where the ToC is located in the sheet, or changes in the range of cells utilised for it, this function should be updated."""
    _keys = [
        cell[0].value for cell in wb_values["Table of Contents & Validation"]["B9:B70"]
    ]
    _values = list(range(9, 71))  # row numbers
    return dict(zip(_keys, _values))


def access_NR_table(
    pyxl_NR, old_sheet_names: list[str] | None = None
) -> tuple[pd.DataFrame, list]:
    """Access names, cell references and coordinates of each name range from the python object containing name ranges.
    Returns a (pandas) DataFrame.
    There were some VSME samples that, after processing in openpyxl, returned wrong sheet references for their name ranges (e.g. "[1]General Information" instead of "General Information");
    To handle this, a list of issues (str explaining where the issue is) is returned alongside the DataFrame."""

    rows: list[tuple[str, str | None, str | None, Shape]] = []
    issues: list[str] = []

    for NR, dn in pyxl_NR.items():
        try:
            destinations = list(dn.destinations)
        except AttributeError:
            issues.append(
                f"Name range '{NR}' has an unreadable destination (value: {dn.attr_text!r}). This name range has been ignored in the migration."
            )
            rows.append((NR, None, None, make_shape(None)))
            continue

        if len(destinations) != 1:
            issues.append(
                f"Name range '{NR}' has {len(destinations)} destinations (expected 1; value: {dn.attr_text!r}). This name range has been ignored in the migration."
            )
            rows.append((NR, None, None, make_shape(None)))
            continue

        sheet, rng = destinations[0]
        if old_sheet_names is not None and sheet not in old_sheet_names:
            issues.append(
                f"Name range '{NR}' refers to sheet '{sheet}' which is not present in the old workbook. This name range has been ignored in the migration."
            )
            sheet = None

        # `rng or None` keeps the cell_ranges column NA-detectable (empty -> None);
        # make_shape independently treats an empty range as a NullShape.
        rows.append((NR, sheet, rng or None, make_shape(rng)))

    df_populated = pd.DataFrame(
        rows, columns=["name_ranges", "sheets", "cell_ranges", "cell_shapes"]
    )
    return df_populated, issues


def access_missingNR_table(missingNR: VersionCollection, version: str) -> pd.DataFrame:
    """
    Access sheets, cell references and coordinates of each missing name range.
    The json data contained in missingNR is hard-coded because certain name ranges were not present in old VSME versions.
    The json is saved in src/migration_tool/data, and must be updated for every new VSME version
    (see validations.py for data structure)
    """

    list_of_sheets = missingNR.get_sheets(version)
    list_of_ranges = missingNR.get_ranges(version)
    list_of_shapes = [make_shape(rng) for rng in list_of_ranges]

    return pd.DataFrame(
        {
            "sheets": list_of_sheets,
            "cell_ranges": list_of_ranges,
            "cell_shapes": list_of_shapes,
        }
    )


def apply_changes_NR(
    df: pd.DataFrame,
    version_cell: str,
    version_cell_new: str,
    NR_changes: dict[str, str],
) -> pd.DataFrame:
    """Change name ranges based on the version of the old workbook, and return a df with the updated name ranges.
    This is necessary because some name ranges have changed from version to version.
    Data comes from NR_changes.json, to be updated each time a NR name changes."""

    df_of_changes_NR = pd.DataFrame(
        {
            "name_ranges": NR_changes[version_cell],
            "name_ranges_new": NR_changes[version_cell_new],
        }
    )

    df = df.merge(df_of_changes_NR, on="name_ranges", how="left")

    for i in range(len(df)):
        if pd.notna(df.loc[i, "name_ranges_new"]):
            df.loc[i, "name_ranges"] = df.loc[i, "name_ranges_new"]

    df = df.drop(columns=["name_ranges_new"])

    return df


def change_wastes(df, mapping) -> list[str]:
    """Change waste categories based on mapping_wastes.pkl provided in src/migration_tool/data/mapping.
    Mapping is based on the changes in waste categories in the new EU Regulation (see more in wastes.xlsx in base dir).
    Returns a list of issues encountered during the change (e.g. old waste category not present in the new Regulation)."""

    old_wastes = (
        df.loc[df["name_ranges"] == "TypeOfWasteAxis", "cell_values"]
        .values[0]
        .first_element_row(double_list=False)
    )
    new_wastes = []
    list_wasteissues: list[str] = []
    for i in old_wastes:
        if i is not None:
            if pd.isna(mapping.loc[mapping["old"] == i, "new"].values[0]):
                new_wastes.append(
                    [
                        f"Waste category {i} not present in new Regulation. Please, see https://eur-lex.europa.eu/legal-content/EN/TXT/PDF/?uri=CELEX:32014D0955"
                    ]
                )
                list_wasteissues.append(
                    f"Waste category --{i}-- not present in new Regulation. Please, see https://eur-lex.europa.eu/legal-content/EN/TXT/PDF/?uri=CELEX:32014D0955"
                )
            else:
                new_wastes.append([mapping.loc[mapping["old"] == i, "new"].values[0]])
    df.loc[df["name_ranges"] == "TypeOfWasteAxis", "cell_values"] = make_value(
        new_wastes
    )

    return list_wasteissues


def clean_NR_with_no_data(df):
    """Function to filter name ranges to process.
    Some (most starting with template_) were created for Digital Converter and not for storing data.
    Unfortunately, there were some exceptions and edge cases (see list_of_NRs_toNOTremove)."""

    list_of_keywords = [
        "template_",
        "enum_",
        "Table",
        "BreakdownOfEnergyConsumptionAxis",
        "Hypercube",
    ]

    list_of_NRs_toNOTremove = [
        "template_reporting_entity_name",
        "template_reporting_entity_identifier_scheme",
        "template_reporting_entity_identifier",
        "template_currency",
    ]
    df_toattach = pd.concat(
        [
            df[df["name_ranges"].isin(list_of_NRs_toNOTremove)],
            df[df["name_ranges"].str.contains("template_checkbox", na=False)],
        ]
    )

    for kw in list_of_keywords:
        df = df[~df["name_ranges"].str.contains(kw, na=False)]

    return pd.concat([df.reset_index(drop=True), df_toattach]).reset_index(drop=True)


def get_indexes_of_NAs(df, column) -> list[int]:
    """Get the indexes of the rows where the specified column has NA values"""
    return df.loc[df[column].isna()].index.tolist()


@overload
def copy_values(pyxl, df, *, NRflag: Literal[True]) -> pd.DataFrame: ...
@overload
def copy_values(pyxl, df, *, NRflag: Literal[False]) -> pd.Series: ...
def copy_values(
    pyxl: Workbook, df: pd.DataFrame, *, NRflag: bool
) -> pd.DataFrame | pd.Series:
    """Copy values from openpyxl workbook based on cell shapes, returns DataFrame with col of cell values.
    2 cases: 1) if NRflag is True, a DataFrame with 2 cols (name ranges and values) is returned;
             2) if NRflag is False, just the cell values are returned as a pandas Series."""

    cell_values = []
    index_sheets = get_indexes_of_NAs(df, "sheets")
    index_ranges = get_indexes_of_NAs(df, "cell_ranges")

    for i in range(len(df)):
        if i not in index_sheets:
            sheet = pyxl[df["sheets"][i]]
            shape = df["cell_shapes"][i]

            # handling the #REF ranges
            if i not in index_ranges:
                cell_values.append(make_value(shape.build_values(sheet)))
            else:
                cell_values.append(make_value(None))
        # to handle issue with sheet names (ex "[1]General Information") in old workbooks
        else:
            cell_values.append(make_value(None))

    df["cell_values"] = cell_values

    if NRflag:
        return df[["name_ranges", "cell_values"]]
    else:
        return df["cell_values"]


def paste_values(pyxl, df, NR=None, table_of_contents=None, version=None):
    """Paste values from a DataFrame column into an openpyxl workbook.

    Treatment for merged cells (merged only across columns in VSMEs):
    - thanks to .merged_cells.ranges method, coordinates of merged cells are extracted and saved in a DataFrame (merged_loc);
    - before pasting, check if top-left cell of the shape is in the merged_loc df, and if so,
      we keep only the first element of each row of values (with .first_element_row method) to fit merged shape.

    Treatment for enlarged ranges (throughout VSME versions) are handled with .enlarged_range_correction method,
    which pastes values only in the original shape of the name range and None(s) for the rest of the enlarged range.

    The NR=True flag argument is used for specifically handling the checkboxes of the biodiversity sites, as if True is pasted the rest of the cols must have False for uniformity (see Template).
    """

    sheet_names = pyxl.sheetnames
    merged_loc = pd.DataFrame()

    for sheet in sheet_names:
        merged = list(pyxl[sheet].merged_cells.ranges)
        merged_list = []
        for i in range(len(merged)):
            merged_list.append(range_boundaries(str(merged[i]))[:2])
        merged_loc = pd.concat([merged_loc, pd.DataFrame({sheet: merged_list})], axis=1)

    add_checkbox = [
        "SiteLocatedInABiodiversitySensitiveArea",
        "SiteLocatedNearABiodiversitySensitiveArea",
    ]

    for i in range(len(df)):
        sheet = pyxl[df["sheets"][i]]
        rng = sheet[df["cell_ranges"][i]]
        shape = df["cell_shapes"][i]
        value = df["cell_values"][i]

        # specific handling for list of classified information
        if table_of_contents is not None:
            if (
                df["name_ranges"][i]
                == "ListOfOmittedDisclosuresDeemedToBeClassifiedOrSensitiveInformation"
            ):
                classified_info_handling(value, table_of_contents, pyxl, version)
                continue

        if value.values() is None:  # nothing to paste (null value or a formula)
            continue

        if shape.isonecell():
            rng.value = value.topleft()  # when it's one cell, paste topleft

        else:
            tuple_to_check = (shape.left(), shape.top())

            if (
                tuple_to_check in merged_loc[df["sheets"][i]].values.tolist()
            ):  # merged cells check
                value.first_element_row()  # keeping only the first value of each row
                value.enlarged_range_correction(shape)  # enlarged ranges check
                value.paste(sheet, shape)

            else:
                value.enlarged_range_correction(shape)  # enlarged ranges check
                if NR is not None:
                    if df["name_ranges"][i] in add_checkbox:
                        value.add_checkboxes()  # add checkboxes for the specific NRs (see above)
                value.paste(sheet, shape)


def classified_info_handling(value, table_of_contents, pyxl, version) -> None:
    """Specific handling for the list of classified information (see new functionality added in version 1.2.0).
    Check where the first 2 characters (ex B1) of "value" (choices about classified info in old versions) match in the ToC,
    and add True in the corresponding row in col D of new "Table of Contents & Validation" sheet where formulas are not detected (formula cells get automatically updated upon opening new version).

    For migrations from 1.2.0 to newer versions, the cells are hard-coded in missingNR_df.
    """

    for input in value.values():
        if input[0] is not None:
            if version in ["1.0.0", "1.0.1", "1.1.0", "1.1.1"]:
                match = [
                    s for s in table_of_contents.keys() if input[0][0:2] in s
                ]  # matches based on first 2 characters

                if match:
                    row_n = [table_of_contents.get(key) for key in match]
                    for row in row_n:
                        cell = pyxl["Table of Contents & Validation"].cell
                        if not check_formula(cell(row=row, column=4)):
                            cell(row=row, column=4).value = True


def adjust_classified_info(
    df: pd.DataFrame, df_values: pd.DataFrame | pd.Series, px_values: Workbook
) -> None:
    """Copy classified info values for version where data is already in cells with formulas.
    Cannot do this in copy_values or build_values method because it needs to be done on value-only workbooks.
    Needed to migrate classified information choices, which in newer templates are selected via the ToC."""

    sheet_generalinfo = px_values["General Information"]
    shape_classifiedinfo = df.loc[
        df["name_ranges"]
        == "ListOfOmittedDisclosuresDeemedToBeClassifiedOrSensitiveInformation",
        "cell_shapes",
    ].item()
    values_classifiedinfo = make_value(
        shape_classifiedinfo.build_values(sheet_generalinfo)
    )

    df_values.loc[
        df["name_ranges"]
        == "ListOfOmittedDisclosuresDeemedToBeClassifiedOrSensitiveInformation",
        "cell_values",
    ] = values_classifiedinfo


def adjust_data_missing_first2versions(
    df_new_wv: pd.DataFrame, missingNR_df_new_wv: pd.DataFrame, old_wb: Workbook
) -> None:

    def add_TrueOrFalse_to_df(df, comb, bool) -> None:
        sheet: str = comb["sheet"]
        rng: str = comb["rng"]
        val = make_value([[bool]])
        df.loc[(df["sheets"] == sheet) & (df["cell_ranges"] == rng), "cell_values"] = (
            val
        )

    # resolving whether undertaking operates in more than one country
    length = (
        df_new_wv.loc[
            df_new_wv["name_ranges"] == "CountryOfEmploymentContractAxis",
            "cell_values",
        ]
        .values[0]
        .count_uniques()
    )
    if length > 2:
        add_TrueOrFalse_to_df(
            missingNR_df_new_wv, {"sheet": "Social Disclosures", "rng": "$E$27"}, True
        )
    else:
        add_TrueOrFalse_to_df(
            missingNR_df_new_wv, {"sheet": "Social Disclosures", "rng": "$E$27"}, False
        )

    # resolving Fuel Converter transfer cell
    first_fuel = old_wb["Fuel Converter"]["B10"].value
    if first_fuel:
        add_TrueOrFalse_to_df(
            missingNR_df_new_wv, {"sheet": "Fuel Converter", "rng": "$D$23"}, True
        )
    else:
        add_TrueOrFalse_to_df(
            missingNR_df_new_wv, {"sheet": "Fuel Converter", "rng": "$D$23"}, False
        )


def create_or_update_migration_status(pyxl) -> None:
    """Create or update the name range template_migration_status in the Introduction sheet, cell D1.
    The cell returns TRUE if the migration process has been completed, but will be processed as None in openpyxl (data_only=True) mode
    if the workbook has not been opened after migration.
    The name range is used as starting check in the Converter, in the case of workbooks of the newest version."""

    if pyxl.defined_names.get("template_migration_status") is None:
        ws = pyxl["Introduction"]
        ref = f"{quote_sheetname(ws.title)}!{absolute_coordinate('D2')}"
        defn = DefinedName(name="template_migration_status", attr_text=ref)

        pyxl.defined_names.add(defn)
        ws["D1"].value = "Migration status"
        ws["D1"].alignment = Alignment(horizontal="center")
        ws["D2"].value = "=AND(TRUE,OR(FALSE,TRUE))"


def assess_energyConsumption_validation(df: pd.DataFrame) -> list[str] | None:
    """Fuel Converter for first 3 versions does not sum from the third added fuel.
    This function checks whether a third fuel (in C12) has been added,
    and returns an issue to be added to the migration issues list since
    the newest version should display a validation error next to the energy cons. cells."""

    if (
        df.loc[
            (df["sheets"] == "Fuel Converter") & (df["cell_ranges"] == "$C$12"),
            "cell_values",
        ]
        .tolist()[0]
        .topleft()
    ):
        return [
            "Issues in Energy Consumption sums. Please check the Environmental Disclosures and the Fuel Converter sheets."
        ]
    else:
        return None
