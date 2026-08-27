import io
import time
import warnings
from collections.abc import Generator
from contextlib import contextmanager
from importlib.resources import as_file, files
from io import BytesIO
from pathlib import Path
from typing import TypeAlias
import json
from dataclasses import dataclass, field

import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook as openpyxl_load_workbook

from .decorators import cached_copy
from .outils import (
    access_missingNR_table,
    access_NR_table,
    add_helperComments_social,
    adjust_classified_info,
    adjust_countriesOfOperation,
    adjust_wasteValues,
    apply_changes_NR,
    assess_energyConsumption_validation,
    change_wastes,
    check_status_incomplete,
    clean_NR_with_no_data,
    convert_months,
    copy_values,
    create_or_update_migration_status,
    create_table_of_contents,
    get_version,
    paste_values,
)
from .data.validations import VersionCollection

FilePathOrBinaryBlob: TypeAlias = str | Path | io.BufferedIOBase
NEW_TEMPLATE_NAME = "VS-Digital-Template-2.0.xlsx"
STATUS_NR = "template_overall_validation_status"


@dataclass
class IssuesCollector:
    issues: list[str] = field(default_factory=list)

    def append_first(self, issue: str):
        self.issues.insert(0, issue)


def load_workbook_quietly(
    file: FilePathOrBinaryBlob, data_only: bool = False, read_only: bool = False
) -> Workbook:
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message=".*? extension is not supported and will be removed",
            category=UserWarning,
            module=r"openpyxl\.worksheet\._reader",
        )
        return openpyxl_load_workbook(file, data_only=data_only, read_only=read_only)


@contextmanager
def _open_cached_values(
    source: FilePathOrBinaryBlob,
) -> Generator[Workbook, None, None]:
    """Open `source` read-only for Excel's cached computed values, and always close it.

    Formula cells surface their cached value here (not the formula). read_only keeps
    this cheap because only a few cells are read.
    """
    wb = load_workbook_quietly(source, data_only=True, read_only=True)
    try:
        yield wb
    finally:
        wb.close()


@cached_copy
def _mapping_wastes() -> pd.DataFrame:
    return pd.read_pickle(
        files("migration_tool.data").joinpath("Mapping_wastes.pkl").open("rb")
    )


@cached_copy
def _missing_NR() -> dict[str, dict[str, str]]:
    with open(
        "src/migration_tool/data/missing_NR.json", "r", encoding="utf-8"
    ) as json_file:
        return json.load(json_file)


@cached_copy
def _NR_changes() -> dict:
    with open(
        "src/migration_tool/data/NR_changes.json", "r", encoding="utf-8"
    ) as json_file:
        return json.load(json_file)


@cached_copy
def _table_of_contents() -> dict[str, int]:
    with (
        as_file(files("migration_tool.data").joinpath(NEW_TEMPLATE_NAME)) as path,
        _open_cached_values(path) as wb,
    ):
        return create_table_of_contents(wb)


def migrate_workbook_as_bytes(
    old_wb: FilePathOrBinaryBlob,
) -> tuple[bytes, float, list[str]]:
    new_wb, elapsed, issues = migrate_workbook(old_wb)
    new_wb_bytes = BytesIO()
    new_wb.save(new_wb_bytes)
    return new_wb_bytes.getvalue(), elapsed, issues


def migrate_workbook(
    old_wb: Workbook | FilePathOrBinaryBlob,
) -> tuple[Workbook, float, list[str]]:
    start_time = time.time()

    mapping_wastes = _mapping_wastes()
    missingNR = VersionCollection().from_dict(_missing_NR())

    # load old filled-out Template (formula-aware; cached values are read separately,
    # read-only, in the `with` block below)
    if isinstance(old_wb, FilePathOrBinaryBlob):
        old_wb_obj = load_workbook_quietly(old_wb, data_only=False)
    else:
        raise TypeError(
            f"old_wb [{type(old_wb)}] must be a file path or an openpyxl Workbook"
        )

    # list of migration issues to return (and to be displayed in webpage after migration).
    # Remember to update if any new potential issues arise.
    c = IssuesCollector()

    # load new empty Template (mutated into the output below, so always fresh)
    with as_file(files("migration_tool.data").joinpath(NEW_TEMPLATE_NAME)) as path:
        new_wb_empty = load_workbook_quietly(path, data_only=False)

    table_of_contents = _table_of_contents()

    version_cell = get_version(old_wb_obj, c)
    version_cell_new = get_version(new_wb_empty, c)

    old_wb_sheet_names = [sheet.title for sheet in old_wb_obj.worksheets]

    df_old = access_NR_table(old_wb_obj.defined_names, c, old_wb_sheet_names)
    df_new = access_NR_table(new_wb_empty.defined_names, c)

    missingNR_df_old = access_missingNR_table(missingNR, version_cell)
    missingNR_df_new = access_missingNR_table(missingNR, version_cell_new)

    df_old_wv = copy_values(old_wb_obj, df_old, NRflag=True)
    missingNR_old_values = copy_values(old_wb_obj, missingNR_df_old, NRflag=False)

    DATES = ("StartDate", "EndDate", "DateAdoptionTransitionPlan")
    convert_months(missingNR_old_values, version_cell, DATES)

    # Read the old workbook's cached computed values once (read-only): the ToC status
    # cell and, for 1.2.0+, the classified-info column (both are formula cells).
    with _open_cached_values(old_wb) as old_values:
        incomplete = check_status_incomplete(old_values, STATUS_NR, c)
        if version_cell not in ["1.0.0", "1.0.1", "1.1.0", "1.1.1"]:
            adjust_classified_info(df_old, df_old_wv, old_values)
        if version_cell in ["1.0.0", "1.0.1", "1.1.0", "1.1.1", "1.2.0", "1.3.0"]:
            c.issues.append(
                "New Template requires explicit statement of compliance (see 'General Information' for more)."
            )
            add_helperComments_social(
                old_values, new_wb_empty, missingNR_df_new, version_cell, c
            )

    # "Incomplete" message first
    if incomplete:
        c.append_first(
            "The old workbook is incomplete. Migration happened only for the filled-out cells, but some data might be missing."
        )

    df_old_tomerge = apply_changes_NR(
        df_old_wv, version_cell, version_cell_new, _NR_changes()
    )
    df_old_tomerge = clean_NR_with_no_data(df_old_tomerge)
    if version_cell in ["1.0.0", "1.0.1"]:
        change_wastes(df_old_tomerge, mapping_wastes, c)

    df_new_wv = df_new.merge(df_old_tomerge)
    missingNR_df_new_wv = missingNR_df_new.merge(missingNR_old_values)

    if version_cell in ["1.0.0", "1.0.1"]:
        adjust_countriesOfOperation(df_new_wv, missingNR_df_new_wv, old_wb_obj)

    if version_cell in ["1.0.0", "1.0.1", "1.1.0"]:
        assess_energyConsumption_validation(missingNR_df_new_wv, c)

    if version_cell in ["1.0.0", "1.0.1", "1.1.0", "1.1.1", "1.2.0", "1.3.0"]:
        df_new_wv = adjust_wasteValues(df_old_tomerge, df_new_wv, new_wb_empty, c)

    paste_values(new_wb_empty, missingNR_df_new_wv)
    paste_values(
        new_wb_empty,
        df_new_wv,
        NR=True,
        table_of_contents=table_of_contents,
        version=version_cell,
    )

    create_or_update_migration_status(new_wb_empty)

    elapsed = time.time() - start_time
    return new_wb_empty, elapsed, c.issues
