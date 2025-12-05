from openpyxl import load_workbook
import pandas as pd
import warnings
import time
import os
from pathlib import Path

from outils.apply_changes_NR import apply_changes_NR, change_wastes
from outils.clean_NR_with_no_data import clean_NR_with_no_data
from outils.paste_values import paste_values
from outils.access_NR_tables import access_NR_table, access_missingNR_table
from outils.copy_values import copy_values
from outils.classes import values

# Resolve paths relative to this script file so pickles are found regardless of current working directory
_base_dir = Path(__file__).resolve().parent
_mapping_dir = _base_dir / "outils"

mapping_wastes = pd.read_pickle(str(_mapping_dir / "Mapping_wastes.pkl"))
missingNR_df = pd.read_pickle(str(_mapping_dir / "missingNR_df.pkl"))


def tool(
    old_wb,
    template_path: str | os.PathLike | None = None,
):
    """
    Run the migration tool on the given Excel file_path.
    Returns:
        output_file (openpyxl.Workbook): new template with old data.
        elapsed (float): execution time in seconds.
    """

    if template_path is None:
        template_path = _base_dir / "outils" / "VSME-Digital-Template-1.1.1.xlsx"
    template_path = str(Path(template_path))

    start_time = time.time()
    # load New (empty) template workbook
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", category=UserWarning)
        new_wb_empty = load_workbook(template_path)

    if not new_wb_empty:
        print("Error: Could not load the new template workbook.")
    else:
        print("New template workbook loaded successfully.")

    # load old workbook if a file path (str or PathLike) was provided, otherwise assume it's already a Workbook
    if isinstance(old_wb, (str, os.PathLike)):
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", category=UserWarning)
            old_wb_obj = load_workbook(old_wb)
    elif old_wb is None:
        raise ValueError("old_wb must be a file path or an openpyxl Workbook")
    else:
        # already a workbook-like object
        old_wb_obj = old_wb

    version_cell = old_wb_obj["Introduction"].cell(row=1, column=3).value
    version_cell_new = new_wb_empty["Introduction"].cell(row=1, column=3).value

    df_old = access_NR_table(old_wb_obj.defined_names)

    df_new = access_NR_table(new_wb_empty.defined_names)

    missingNR_df_old = access_missingNR_table(missingNR_df, version_cell)
    missingNR_df_new = access_missingNR_table(missingNR_df, version_cell_new)

    df_old_wv = copy_values(old_wb_obj, df_old, key="name_ranges")
    missingNR_df_old_values = copy_values(old_wb_obj, missingNR_df_old, key=None)

    if version_cell == "1.0.0":
        for position in [0, 1, 6]:
            missingNR_df_old_values[position].convert_month_to_numbers()

    df_old_tomerge = apply_changes_NR(df_old_wv, version_cell, version_cell_new)
    df_old_tomerge = clean_NR_with_no_data(df_old_tomerge)
    if version_cell in ["1.0.0", "1.0.1", "1.1.0"]:
        change_wastes(df_old_tomerge, mapping_wastes)

    df_new_wv = df_new.merge(df_old_tomerge)
    missingNR_df_new_wv = pd.concat([missingNR_df_new, missingNR_df_old_values], axis=1)

    if version_cell in ["1.0.0", "1.0.1"]:
        length = (
            df_new_wv.loc[
                df_new_wv["name_ranges"] == "CountryOfEmploymentContractAxis",
                "cell_values",
            ]
            .values[0]
            .count_uniques()
        )
        if length > 2:
            missingNR_df_new_wv.loc[
                (missingNR_df_new_wv["sheets"] == "Social Disclosures")
                & (missingNR_df_new_wv["cell_ranges"] == "$E$27"),
                "cell_values",
            ] = values([[True]])
        else:
            missingNR_df_new_wv.loc[
                (missingNR_df_new_wv["sheets"] == "Social Disclosures")
                & (missingNR_df_new_wv["cell_ranges"] == "$E$27"),
                "cell_values",
            ] = values([[False]])

    paste_values(new_wb_empty, missingNR_df_new_wv)
    paste_values(new_wb_empty, df_new_wv, NR=True)

    elapsed = time.time() - start_time
    return new_wb_empty, elapsed


if __name__ == "__main__":
    # simple runner when executed as a script
    tool(
        "Template/VSME-Digital-Template-Sample-1.0.0.xlsx",
        "Template/VSME-Digital-Template-1.1.1.xlsx",
    )
